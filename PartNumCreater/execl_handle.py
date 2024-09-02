import pandas as pd
import copy
import os
from datetime import datetime

data = {}
list_data = []

def excel_file_read(file_name: str, sheet_name: str):
    # 讀取 Excel 文件
    return pd.read_excel(file_name, sheet_name)

def excel_file_single_adjust(file_path, sheet_name, row, column_name):
    # 保存修改后的数据到新的 Excel 文件
    try:
        # 加载现有的工作簿
        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        # 获取指定的工作表
        ws = wb[sheet_name]

        # 读取单元格并修改内容
        print(row)
        cell_value = int(ws["E"+str(row+2)].value)  # 注意行索引需要加2，因为 openpyxl 是以1开始的而不是0
        ws["E"+str(row+2)] = cell_value + 1

        # 保存修改后的文件
        wb.save(file_path)
        print(f"修改后的数据已保存到 {file_path} 的工作表 {sheet_name} 中。")

    except Exception as e:
        print(f"保存 Excel 文件失败: {e}")



#get KEY array
def get_key_array(data: dict):
    unique_items = set()
    # Extract and add items from 'SMT'
    for key in data.values():
        for item in key:
            unique_items.update(item.keys())

    # Convert the set to a list
    unique_list = list(unique_items)
    return unique_list

#return capacity
def add_capacity(df2):
    return_dict = {}
    X7R_list = []
    X5R_list = []
    NPO_list = []


    X7R_column_values_code = df2.iloc[:, 7]
    X5R_column_values_code = df2.iloc[:, 9]
    NPO_column_values_code = df2.iloc[:, 11]

    for index in range(X7R_column_values_code.size):
        if (X7R_column_values_code[index] != "4.編號"  and pd.notna(X7R_column_values_code[index])):
            X7R_list.append(str(X7R_column_values_code[index]))
            X5R_list.append(X5R_column_values_code[index])
            NPO_list.append(NPO_column_values_code[index])

    return_dict.update({'X7R':X7R_list})
    return_dict.update({'X5R':X5R_list})
    return_dict.update({'NPO':NPO_list})

    return return_dict

#return array
def excel_list_read(df):
    return_value = []

    column_values = df.iloc[:, 8]
    column_values_code = df.iloc[:, 7]

    for len in range(column_values.size):
        if (column_values[len] == "項目"):
            if ({column_values[len+1]:column_values_code[len+1]} not in return_value):
                return_value.append({column_values[len+1]:column_values_code[len+1]})
    
    return return_value

def excel_type_read(df):
    type_value = []
    key_list = []
    return_value = {}
    list_data = ""

    column_values = df.iloc[:, 10]
    column_values_code = df.iloc[:, 9]
    column_values_list = df.iloc[:, 8]

    for len in range(column_values.size):
        if (column_values_list[len] == "項目"):
            if(list_data != column_values_list[len+1]):
                list_data = column_values_list[len+1]
                type_value.clear()
        if (column_values[len] == "種類"):
            if ({column_values[len+1].strip():str(column_values_code[len+1])} not in type_value):
                type_value.append({column_values[len+1].strip():str(column_values_code[len+1])})
                return_value.update({list_data:copy.deepcopy(type_value)})

    key_list = get_key_array(return_value)

    return key_list, return_value

def excel_type_read_supplier(df):
    column_values = df.iloc[:, 1]
    company_dict = [company.split()[0] for company in column_values]

    return company_dict
 
def execl_size_read(df, key_list: list):
    size_value = []
    storage_list = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 12]
    column_values_code = df.iloc[:, 11]
    column_values_type = df.iloc[:, 10]

    for length in range(column_values.size):
        if (column_values_type[length] == "種類"):
            if(type_data != column_values_type[length+1]):
                type_data = column_values_type[length+1]

        if (column_values[length] != "零件尺寸" and column_values[length] != "種類" and column_values[length] != "間距" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])
    
    for item in key_list:
        size_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                if ({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])} not in size_value):
                    size_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                    return_value.update({item:copy.deepcopy(size_value)})

    key_list = get_key_array(return_value)

    print(return_value)
        
    return key_list, return_value

def execl_percentage_read(df, key_list: list):
    percentage_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 14]
    column_values_code = df.iloc[:, 13]
    column_values_type = df.iloc[:, 10]
    print(column_values)
    for length in range(column_values.size):
        if (column_values_type[length] == "種類"):
            if(type_data != column_values_type[length+1]):
                type_data = column_values_type[length+1]

        if (column_values[length] != "%數" and column_values[length] != "封裝" and column_values[length] != "開關位數"  and column_values[length] != "預留" and column_values[length] != "總PIN數" and pd.notna(column_values[length])):
            
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        percentage_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                percentage_value.append({str(column_values[storage_list[i][2]]).strip().replace('\u3000','').replace(' ',''):str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(percentage_value)})

    return return_value

def execl_capacity_read(df, key_list: list):
    capacity_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 16]
    column_values_code = df.iloc[:, 15]
    column_values_type = df.iloc[:, 10]

    for length in range(column_values.size):
        if (column_values_type[length] == "種類"):
            if(type_data != column_values_type[length+1]):
                type_data = column_values_type[length+1]

        if (column_values[length] != "電容值" and column_values[length] != "電阻值" and column_values[length] != "PCB名稱" and column_values[length] != "零件名稱" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        capacity_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                capacity_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(capacity_value)})

    return return_value

def execl_packaging_read(df, key_list: list):
    capacity_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 14]
    column_values_code = df.iloc[:, 13]
    column_values_kind = df.iloc[:, 12]
    column_values_odject = df.iloc[:, 8]

    for length in range(column_values.size):
        if (column_values_kind[length] == "種類" or column_values_kind[length] == "間距"):
            if(type_data != column_values_kind[length+1]):
                type_data = column_values_kind[length+1]

        if (column_values[length] != "%數" and column_values[length] != "封裝"  and column_values[length] != "開關位數" and column_values[length] != "總PIN數" and column_values[length] != "大版本" and column_values[length] != "電流" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        capacity_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                capacity_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(capacity_value)})

    return return_value

def execl_name_read(df, key_list: list):
    capacity_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 16]
    column_values_code = df.iloc[:, 15]
    column_values_kind = df.iloc[:, 12]

    for length in range(column_values.size):
        if (column_values_kind[length] == "種類" or column_values_kind[length] == "間距"):
            if(type_data != column_values_kind[length+1]):
                type_data = column_values_kind[length+1]

        if (type_data != '' and column_values[length] != "電容值" and column_values[length] != "電阻值" and column_values[length] != "PCB名稱" and column_values[length] != "零件名稱" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        capacity_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                capacity_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(capacity_value)})

    return return_value

def execl_voltage_read(df, key_list: list):
    voltage_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 18]
    column_values_code = df.iloc[:, 17]
    column_values_type = df.iloc[:, 10]

    for length in range(column_values.size):
        if (column_values_type[length] == "種類"):
            if(type_data != column_values_type[length+1]):
                type_data = column_values_type[length+1]

        if (column_values[length] != "電壓" and column_values[length] != "腳位大小" and column_values[length] != "顏色" and column_values[length] != "頻率" and column_values[length] != "小版本" and column_values[length] != "版本" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        voltage_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                voltage_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(voltage_value)})

    return return_value

def execl_manufacturer_read(df, key_list: list):
    manufacturer_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 20]
    column_values_code = df.iloc[:, 19]
    column_values_type = df.iloc[:, 10]

    for length in range(column_values.size):
        if (column_values_type[length] == "種類"):
            if(type_data != column_values_type[length+1]):
                type_data = column_values_type[length+1]

        if (column_values[length] != "廠商" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        manufacturer_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                manufacturer_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(manufacturer_value)})

    return return_value

def execl_manufacturer_2_read(df, key_list: list):
    manufacturer_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 20]
    column_values_code = df.iloc[:, 19]
    column_values_kind = df.iloc[:, 12]

    for length in range(column_values.size):
        if (column_values_kind[length] == "種類" or column_values_kind[length] == "間距"):
            if(type_data != column_values_kind[length+1]):
                type_data = column_values_kind[length+1]

        if (column_values[length] != "廠商" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        manufacturer_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                manufacturer_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(manufacturer_value)})

    return return_value

def execl_supplier_read(df, key_list: list):
    manufacturer_value = []
    storage_list = []
    return_value = {}
    type_data = ""

    column_values = df.iloc[:, 22]
    column_values_code = df.iloc[:, 21]
    column_values_type = df.iloc[:, 10]

    for length in range(column_values.size):
        if (column_values_type[length] == "種類"):
            if(type_data != column_values_type[length+1]):
                type_data = column_values_type[length+1]

        if (column_values[length] != "供應商" and pd.notna(column_values[length])):
            storage_list.append([type_data,column_values[length],length])

    for item in key_list:
        manufacturer_value.clear()
        for i in range(len(storage_list)):
            if (storage_list[i][0].strip() == item.strip()):
                manufacturer_value.append({str(column_values[storage_list[i][2]]).strip():str(column_values_code[storage_list[i][2]])})
                return_value.update({item:copy.deepcopy(manufacturer_value)})

    return return_value

def execl_mechanism_element_read(df):
    hundreds_list = [{row['Unnamed: 1']: str(int(row['使用PIN數']))} for _, row in df.iterrows() if not pd.isna(row['使用PIN數'])]
    tens_list = [{row['Unnamed: 3']: str(int(row['角度']))} for _, row in df.iterrows() if not pd.isna(row['角度'])]
    units_list = [{row['Unnamed: 5']: str(row['顏色'])} for _, row in df.iterrows() if not pd.isna(row['顏色'])]

    result = {
        '使用PIN數': hundreds_list,
        '角度': tens_list,
        '顏色': units_list
    }
    return result

def check_output_existing(file_path):
    try:
        existing_df = pd.read_excel(file_path)

        return existing_df
    except:
        return None

def append_data_to_excel(file_path, part_number, values, headers):
    try:
        existing_df = pd.read_excel(file_path)
    except FileNotFoundError:
        existing_df = pd.DataFrame(columns=headers)


    add_row = [part_number] + values
    new_row = pd.DataFrame([add_row], columns=headers)
    
    existing_df = pd.concat([existing_df, new_row], ignore_index=True)

    try:
        if len(headers) != len(existing_df.columns):
      
            raise ValueError("Headers length does not match the number of columns in the DataFrame.")

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            existing_df.to_excel(writer, index=False, header=headers)
    except Exception as e:
        from component import show_alert
        show_alert("儲存物料["+part_number+"]時發生錯誤\n+"+str(e))
        return False

    return True

def find_item_in_stock(part_number, in_stock_date, stock_place):
    try:
        existing_df = pd.read_excel("stock.xlsx")
    except FileNotFoundError:
        return False

    matching_rows = existing_df[(existing_df['物料編碼'] == part_number) &
                   (existing_df['入庫日期'] == in_stock_date) &
                   (existing_df['倉庫'] == stock_place)]
    
    return existing_df, matching_rows

def execl_stock_record(part_number, in_stock_date, count, total_count, stock_place, recorder, check_code, line_bar2):
    if (check_code == 0):
        try:
            existing_df = pd.read_excel("stock.xlsx")
        except FileNotFoundError:
            existing_df = pd.DataFrame(columns=["物料編碼","入庫日期","總數量","倉庫","記錄人"])
        
        add_row = [part_number, in_stock_date, total_count, stock_place,recorder] 
        new_row = pd.DataFrame([add_row], columns=["物料編碼","入庫日期","總數量","倉庫","記錄人"])

        existing_df = pd.concat([existing_df, new_row], ignore_index=True)

        try:
            with pd.ExcelWriter("stock.xlsx", engine='openpyxl') as writer:
                existing_df.to_excel(writer, index=False, header=["物料編碼","入庫日期","總數量","倉庫","記錄人"])

            # 使用 openpyxl 设置密码保护
            import openpyxl 

            # 加载保存的数据
            wb = openpyxl.load_workbook("stock.xlsx")

            worksheet = wb.active
    
            # 设置工作簿保护
            worksheet.protection.set_password('77777777')
            worksheet.protection.enable()

            wb.save("stock.xlsx")
            wb.close()
                
        except:
            from component import show_alert
            show_alert("儲存物料["+part_number+"]時發生錯誤\nstock.xlxs已被使用")
            return False

        return True
    
    elif (check_code == 1):
        existing_df, matching_rows = find_item_in_stock(part_number, in_stock_date, stock_place)
        
        if not matching_rows.empty:
            row_index = matching_rows.index[0]  # 假设只有一行匹配
            print(count)
            new_total = count
            # if new_total > 0:
            existing_df.at[row_index, '總數量'] = new_total
            # else:
            #     existing_df.drop(row_index, inplace=True)
            #     print(f"Row with material code '{part_number}', date '{in_stock_date}', and warehouse '{stock_place}' has been removed due to zero or negative total quantity.")
            existing_df.to_excel('stock.xlsx', index=False)
            # 使用 openpyxl 设置密码保护
            import openpyxl 

            # 加载保存的数据
            wb = openpyxl.load_workbook("stock.xlsx")

            worksheet = wb.active
    
            # 设置工作簿保护
            worksheet.protection.set_password('77777777')
            worksheet.protection.enable()

            wb.save("stock.xlsx")
            wb.close()

        else:
            print("No matching row found")
        
        

    elif (check_code == 2):
        existing_df, matching_rows = find_item_in_stock(part_number, in_stock_date, stock_place)
        
        # 确认是否找到匹配的行
        if not matching_rows.empty:
            row_index = matching_rows.index[0]  # 假设只有一行匹配
            new_total = count
            # if new_total > 0:
            existing_df.at[row_index, '總數量'] = new_total
            existing_df.to_excel('stock.xlsx', index=False)
            # 使用 openpyxl 设置密码保护
            import openpyxl 

            # 加载保存的数据
            wb = openpyxl.load_workbook("stock.xlsx")

            worksheet = wb.active
    
            # 设置工作簿保护
            worksheet.protection.set_password('77777777')
            worksheet.protection.enable()

            wb.save("stock.xlsx")
            wb.close()
        else:
            print("No matching row found")
    
    line_bar2.setText(str(new_total))

def export_to_inventory_table(inventory_id, erp_code, total_quantity, location, entry_date, edit_by, file_name='stock.xlsx'):
    columns = ['InventoryID', 'ERP_code', 'TotalQuantity', 'Location', 'EntryDate', 'Editby']

    # 使用 openpyxl 设置密码保护
    import openpyxl 

    # 检查文件和工作表是否存在
    if os.path.exists(file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            if 'Inventory' in workbook.sheetnames:
                df = pd.read_excel(file_name, sheet_name='Inventory')
            else:
                df = pd.DataFrame(columns=columns)
        except Exception as e:
            print(f"Error loading workbook: {e}")
            df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(columns=columns)

    # 检查inventory_id是否已存在
    if inventory_id in df['InventoryID'].values:
        df.loc[df['InventoryID'] == inventory_id, 'TotalQuantity'] += total_quantity
    else:
        new_data = pd.DataFrame({
            'InventoryID': [inventory_id],
            'ERP_code': [erp_code],
            'TotalQuantity': [total_quantity],
            'Location': [location],
            'EntryDate': [entry_date],
            'Editby': [edit_by]
        })

        # 将新数据附加到现有数据
        df = pd.concat([df, new_data], ignore_index=True)

    # 保存到 Excel 文件
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
        if 'Inventory' in writer.book.sheetnames:
            workbook = writer.book
            workbook.remove(workbook['Inventory'])
            df.to_excel(writer, sheet_name='Inventory', index=False)
        else:
            df.to_excel(writer, sheet_name='Inventory', index=False)

    # 加载保存的数据
    wb = openpyxl.load_workbook("stock.xlsx")

    worksheet = wb.active

    # 设置工作簿保护
    worksheet.protection.set_password('77777777')
    worksheet.protection.enable()

    wb.save("stock.xlsx")
    wb.close()

def number_to_letter(num):
    if num < 10:
        return chr(ord('0') + num)  # 0-9 轉換為 '0'-'9'
    elif num < 36:
        return chr(ord('A') + num - 10)  # 10-35 轉換為 'A'-'Z'
    elif num < 62:
        return chr(ord('a') + num - 36)  # 36-61 轉換為 'a'-'z'
    else:
        return None  # 處理超出範圍的情況

def letter_to_number(letter):
    if isinstance(letter, str) and len(letter) == 1:
        if letter.isdigit():
            return letter  # 如果是數字0~9，直接返回字串
        elif letter.isalpha():
            num = ord(letter)
            if 'a' <= letter <= 'z':
                return num - ord('a') + 36
            else:
                return num - ord('A') + 10
    return None  # 處理非字母或非單個字符的情況

def get_box_num(inventory_id, file_name='stock.xlsx'):
    import openpyxl

    if os.path.exists(file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            if 'Boxes' in workbook.sheetnames:
                df = pd.read_excel(file_name, sheet_name='Boxes')
                if inventory_id in df['InventoryID'].values:
                    existing_numbers = df[df['InventoryID'] == inventory_id]['Number']
                    last_number = existing_numbers.max()
                    return last_number+1
                else:
                    return 1
            else:
                return 1
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return 1
    else:
        return 1

def export_to_box_table(box_id ,inventory_id, number, quantity, print_count , file_name='stock.xlsx'):     
    columns = ['BoxID', 'InventoryID', 'Number', 'Quantity', 'Print_Count']

    # 使用 openpyxl 设置密码保护
    import openpyxl 

    # 检查文件和工作表是否存在
    if os.path.exists(file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            if 'Boxes' in workbook.sheetnames:
                df = pd.read_excel(file_name, sheet_name='Boxes')
            else:
                df = pd.DataFrame(columns=columns)
        except Exception as e:
            print(f"Error loading workbook: {e}")
            df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(columns=columns)

    # 检查InventoryID是否已存在，并计算新的Number
    # if inventory_id in df['InventoryID'].values:
    #     existing_numbers = df[df['InventoryID'] == inventory_id]['Number']
    #     last_number = existing_numbers.max()
    #     new_number = last_number + 1
    # else:
    #     new_number = '0'

    # 添加新数据
    new_data = pd.DataFrame({
        'BoxID': [box_id],
        'InventoryID': [inventory_id],
        'Number': [str(number)],
        'Quantity': [quantity],
        'Print_Count': [print_count],
    })

    # 将新数据附加到现有数据
    df = pd.concat([df, new_data], ignore_index=True)

    # 保存到 Excel 文件
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
        if 'Boxes' in writer.book.sheetnames:
            workbook = writer.book
            workbook.remove(workbook['Boxes'])
            df.to_excel(writer, sheet_name='Boxes', index=False)
        else:
            df.to_excel(writer, sheet_name='Boxes', index=False)

    # 加载保存的数据
    wb = openpyxl.load_workbook("stock.xlsx")

    worksheet = wb.active

    # 设置工作簿保护
    worksheet.protection.set_password('77777777')
    worksheet.protection.enable()

    wb.save("stock.xlsx")
    wb.close()

def export_to_operateion_table(box_id, status, quantity_old, quantity_new, edit_by, file_name='stock.xlsx'):
    columns=['OpID', 'BoxID', 'Status', 'Quantity_Old', 'Quantity_New', 'LastUpdated', 'Editby']

    # 使用 openpyxl 设置密码保护
    import openpyxl 

    if os.path.exists(file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            if 'Operation' in workbook.sheetnames:
                df = pd.read_excel(file_name, sheet_name='Operation')
                next_id = df['OpID'].max() + 1 if not df.empty else 1
            else:
                df = pd.DataFrame(columns=columns)
                next_id = 1
        except Exception as e:
            print(f"Error loading workbook: {e}")
            df = pd.DataFrame(columns=columns)
            next_id = 1
    else:
        df = pd.DataFrame(columns=columns)
        next_id = 1

    # 添加新数据
    new_data = pd.DataFrame({
        'OpID': [next_id],
        'BoxID': [box_id],
        'Status': [status],
        'Quantity_Old': [quantity_old],
        'Quantity_New': [quantity_new],
        'LastUpdated': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        'Editby': [edit_by]
    })

    # 将新数据附加到现有数据
    df = pd.concat([df, new_data], ignore_index=True)

    # 保存到 Excel 文件
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
        if 'Operation' in writer.book.sheetnames:
            workbook = writer.book
            workbook.remove(workbook['Operation'])
            df.to_excel(writer, sheet_name='Operation', index=False)
        else:
            df.to_excel(writer, sheet_name='Operation', index=False)

def get_total_quantity(inventory_id, file_name='stock.xlsx'):
    try:
        # 读取Excel文件中的'Inventory'表
        df = pd.read_excel(file_name, sheet_name='Inventory')
        
        # 使用InventoryID过滤行并获取TotalQuantity值
        total_quantity = df.loc[df['InventoryID'] == inventory_id, 'TotalQuantity'].values
        
        if len(total_quantity) > 0:
            return total_quantity[0]
        else:
            return None
    except Exception as e:
        print(f"Error reading the file or sheet: {e}")
        return None

def get_box_quantity(box_id, file_name='stock.xlsx'):
    try:
        # 读取Excel文件中的'Boxes'表
        df = pd.read_excel(file_name, sheet_name='Boxes')
        
        # 使用BoxID过滤行并获取Quantity值
        quantity = df.loc[df['BoxID'] == box_id, 'Quantity'].values
        
        if len(quantity) > 0:
            return quantity[0]
        else:
            return None
    except Exception as e:
        print(f"Error reading the file or sheet: {e}")
        return None

def out_update_quantity(box_id, quantity_to_take, file_name='stock.xlsx'):
    try:
        # 读取Excel文件中的'Boxes'表
        df = pd.read_excel(file_name, sheet_name='Boxes')
        
        # 查找指定BoxID的行
        box_row = df.loc[df['BoxID'] == box_id]
        
        if box_row.empty:
            return f"BoxID {box_id} not found."
        
        # 获取当前的Quantity值
        current_quantity = box_row['Quantity'].values[0]
        
        # 检查取出的数量是否小于当前Quantity
        if quantity_to_take > current_quantity:
            return f"Quantity to take ({quantity_to_take}) is greater than current quantity ({current_quantity})."
        
        # 更新Quantity值
        new_quantity = current_quantity - quantity_to_take
        df.loc[df['BoxID'] == box_id, 'Quantity'] = new_quantity
        
        # 保存更新后的DataFrame到Excel文件
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
            if 'Boxes' in writer.book.sheetnames:
                workbook = writer.book
                workbook.remove(workbook['Boxes'])
                df.to_excel(writer, sheet_name='Boxes', index=False)
            else:
                df.to_excel(writer, sheet_name='Boxes', index=False)

        # 读取Excel文件中的'Boxes'表
        df = pd.read_excel(file_name, sheet_name='Inventory')
        
        # 查找指定BoxID的行
        Inventory_row = df.loc[df['InventoryID'] == box_id[:-1]]
        
        if Inventory_row.empty:
            return f"BoxID {box_id[:-1]} not found."
        
        # 获取当前的Quantity值
        current_total_quantity = Inventory_row['TotalQuantity'].values[0]
        
        # 更新Quantity值
        new_total_quantity = current_total_quantity - quantity_to_take
        df.loc[df['InventoryID'] == box_id[:-1], 'TotalQuantity'] = new_total_quantity
        
        # 保存更新后的DataFrame到Excel文件
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
            if 'Inventory' in writer.book.sheetnames:
                workbook = writer.book
                workbook.remove(workbook['Inventory'])
                df.to_excel(writer, sheet_name='Inventory', index=False)
            else:
                df.to_excel(writer, sheet_name='Inventory', index=False)


        
        
        return [new_quantity, new_total_quantity]
    
    except Exception as e:
        return f"Error: {e}"
    
def back_update_quantity(box_id, quantity_to_take, file_name='stock.xlsx'):
    try:
        # 读取Excel文件中的'Boxes'表
        df = pd.read_excel(file_name, sheet_name='Boxes')
        
        # 查找指定BoxID的行
        box_row = df.loc[df['BoxID'] == box_id]
        
        if box_row.empty:
            return f"BoxID {box_id} not found."
        
        # 获取当前的Quantity值
        current_quantity = box_row['Quantity'].values[0]
        
        # 检查取出的数量是否小于当前Quantity
        # if quantity_to_take > current_quantity:
        #     return f"Quantity to take ({quantity_to_take}) is greater than current quantity ({current_quantity})."
        
        # 更新Quantity值
        new_quantity = current_quantity + quantity_to_take
        df.loc[df['BoxID'] == box_id, 'Quantity'] = new_quantity
        
        # 保存更新后的DataFrame到Excel文件
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
            if 'Boxes' in writer.book.sheetnames:
                workbook = writer.book
                workbook.remove(workbook['Boxes'])
                df.to_excel(writer, sheet_name='Boxes', index=False)
            else:
                df.to_excel(writer, sheet_name='Boxes', index=False)

        # 读取Excel文件中的'Boxes'表
        df = pd.read_excel(file_name, sheet_name='Inventory')
        
        # 查找指定BoxID的行
        Inventory_row = df.loc[df['InventoryID'] == box_id[:-1]]
        
        if Inventory_row.empty:
            return f"BoxID {box_id[:-1]} not found."
        
        # 获取当前的Quantity值
        current_total_quantity = Inventory_row['TotalQuantity'].values[0]
        
        # 更新Quantity值
        new_total_quantity = current_total_quantity + quantity_to_take
        df.loc[df['InventoryID'] == box_id[:-1], 'TotalQuantity'] = new_total_quantity
        
        # 保存更新后的DataFrame到Excel文件
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a' if os.path.exists(file_name) else 'w') as writer:
            if 'Inventory' in writer.book.sheetnames:
                workbook = writer.book
                workbook.remove(workbook['Inventory'])
                df.to_excel(writer, sheet_name='Inventory', index=False)
            else:
                df.to_excel(writer, sheet_name='Inventory', index=False)


        
        
        return [new_quantity, new_total_quantity]
    
    except Exception as e:
        return f"Error: {e}"

def read_excel_sheets(file_path):
    xls = pd.ExcelFile(file_path)
    return xls.sheet_names
    # for sheet_name in sheet_names:
    #     df = pd.read_excel(file_path, sheet_name=sheet_name)
    #     print(f"Sheet name: {sheet_name}")
    
def read_BOM(df):
    data_list = []

    ERP_CODE_LIST = df.iloc[:, 1]
    COUNT_LIST = df.iloc[:, 6]

    for i in range(1, len(ERP_CODE_LIST)):
        data_list.append({ERP_CODE_LIST[i]:COUNT_LIST[i]})

    return data_list

def read_stock_remand(df):
    data_list = []

    ERP_CODE_LIST = df.iloc[:, 1]
    COUNT_LIST = df.iloc[:, 2]

    for i in range(1, len(ERP_CODE_LIST)):
        data_list.append({ERP_CODE_LIST[i]:COUNT_LIST[i]})

    return data_list

# input data and BoxID to search operation
def read_operation_sheet(df, BoxID):
    pass

def export_data_to_excel(data, file_path, header=None, sheet_name="1"):
    

    if (header == None):
        data = [row[:-1] for row in data]
        df = pd.DataFrame(data, columns=[
            'ERP Code', '品項編號', '品項名稱', '項目', '種類',
            '尺寸/種類', '%數/封裝', '容值/阻值/名稱',
            '電壓/腳位大小/頻率', '廠商', '供應商', '產生時間', 'Part Number'
        ])
    else:
        df = pd.DataFrame(data, columns=header)
    try:
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    except Exception as e:
        return e
    else:
        return True



# df = excel_file_read('曜璿東命名規則 20240605-2.xlsx', '命名規則')
# df2 = excel_file_read('曜璿東命名規則 20240605-2.xlsx', '電容種類規則')

# add_capacity(df2)

# print(excel_list_read(df))
# list_data, data = excel_type_read(df)
# print(list_data)
# print(execl_size_read(df, list_data))
# print(execl_percentage_read(df, list_data))
# print(execl_capacity_read(df, list_data))
# execl_voltage_read(df, list_data)
# execl_manufacturer_read(df, list_data)
# print(execl_supplier_read(df, list_data))

# df = excel_file_read('曜璿東命名規則 20240605-2.xlsx', '命名規則')
# df_supplier = excel_file_read('曜璿東命名規則 20240605-2.xlsx', '供應商編碼')
# excel_type_read_supplier(df_supplier)
# export_to_inventory_table(1, 'ERP123', 100, 'Warehouse 1', '2024-07-06', 'UserA')
# export_to_inventory_table(2, 'ERP124', 200, 'Warehouse 2', '2024-07-07', 'UserB')
# export_to_inventory_table(1, 'ERP123', 50, 'Warehouse 1', '2024-07-08', 'UserC')